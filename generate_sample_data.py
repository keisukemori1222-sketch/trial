"""サンプルの業績管理表（Excel）を生成するスクリプト"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import random

random.seed(42)


def generate_sample_excel(filepath: str = "data/業績管理表.xlsx") -> None:
    wb = openpyxl.Workbook()

    # --- シート1: 月次PL（全社） ---
    ws_pl = wb.active
    ws_pl.title = "月次PL"

    months = [f"2025/{m}月" for m in range(4, 13)] + [f"2026/{m}月" for m in range(1, 4)]
    pl_items = [
        "売上高",
        "売上原価",
        "売上総利益（粗利）",
        "販売費及び一般管理費",
        "  人件費",
        "  広告宣伝費",
        "  地代家賃",
        "  減価償却費",
        "  その他販管費",
        "営業利益",
    ]

    # ヘッダー
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws_pl.cell(row=1, column=1, value="勘定科目").font = header_font
    ws_pl.cell(row=1, column=1).fill = header_fill
    ws_pl.cell(row=1, column=1).border = thin_border
    ws_pl.column_dimensions["A"].width = 28

    for j, month in enumerate(months, start=2):
        cell = ws_pl.cell(row=1, column=j, value=month)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws_pl.column_dimensions[get_column_letter(j)].width = 14

    # データ生成（集計行は後から計算するため2パスで処理）
    base_revenue = 50000  # 5億円ベース（千円単位）

    # ヘッダー列と行名を先に書く
    for i, item in enumerate(pl_items, start=2):
        ws_pl.cell(row=i, column=1, value=item).border = thin_border
        if item in ("売上総利益（粗利）", "営業利益"):
            ws_pl.cell(row=i, column=1).font = Font(bold=True)

    # パス1: 個別項目を生成
    for j, _ in enumerate(months, start=2):
        seasonal = 1.0 + 0.15 * ((j - 2) % 12 in [2, 5, 8, 11])  # 四半期末は+15%
        growth = 1.0 + 0.02 * (j - 2)  # 月2%成長

        # 売上高 (row 2)
        val = int(base_revenue * seasonal * growth * (1 + random.uniform(-0.05, 0.05)))
        ws_pl.cell(row=2, column=j, value=val)

        # 売上原価 (row 3)
        val = int(ws_pl.cell(row=2, column=j).value * random.uniform(0.55, 0.62))
        ws_pl.cell(row=3, column=j, value=val)

        # 売上総利益 (row 4) = 売上高 - 売上原価
        ws_pl.cell(row=4, column=j, value=ws_pl.cell(row=2, column=j).value - ws_pl.cell(row=3, column=j).value)

        # 販管費内訳 (rows 6-10)
        ws_pl.cell(row=6, column=j, value=int(base_revenue * 0.12 * growth * (1 + random.uniform(-0.02, 0.02))))  # 人件費
        ws_pl.cell(row=7, column=j, value=int(base_revenue * 0.05 * seasonal * (1 + random.uniform(-0.1, 0.1))))  # 広告宣伝費
        ws_pl.cell(row=8, column=j, value=int(base_revenue * 0.03))  # 地代家賃
        ws_pl.cell(row=9, column=j, value=int(base_revenue * 0.02))  # 減価償却費
        ws_pl.cell(row=10, column=j, value=int(base_revenue * 0.03 * (1 + random.uniform(-0.05, 0.05))))  # その他販管費

        # 販管費合計 (row 5) = sum(rows 6-10)
        ws_pl.cell(row=5, column=j, value=sum(ws_pl.cell(row=r, column=j).value for r in range(6, 11)))

        # 営業利益 (row 11) = 粗利 - 販管費
        ws_pl.cell(row=11, column=j, value=ws_pl.cell(row=4, column=j).value - ws_pl.cell(row=5, column=j).value)

    # セルのスタイルを適用
    for i in range(2, len(pl_items) + 2):
        for j in range(2, 2 + len(months)):
            cell = ws_pl.cell(row=i, column=j)
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
            cell.border = thin_border

    # --- シート2: 部門別売上 ---
    ws_dept = wb.create_sheet("部門別売上")
    departments = ["営業1部", "営業2部", "営業3部", "EC事業部", "コンサル事業部"]

    ws_dept.cell(row=1, column=1, value="部門").font = header_font
    ws_dept.cell(row=1, column=1).fill = header_fill
    ws_dept.cell(row=1, column=1).border = thin_border
    ws_dept.column_dimensions["A"].width = 18

    for j, month in enumerate(months, start=2):
        cell = ws_dept.cell(row=1, column=j, value=month)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws_dept.column_dimensions[get_column_letter(j)].width = 14

    dept_shares = [0.30, 0.25, 0.20, 0.15, 0.10]
    for i, (dept, share) in enumerate(zip(departments, dept_shares), start=2):
        ws_dept.cell(row=i, column=1, value=dept).border = thin_border
        for j in range(2, 2 + len(months)):
            total_revenue = ws_pl.cell(row=2, column=j).value
            val = int(total_revenue * share * (1 + random.uniform(-0.08, 0.08)))
            cell = ws_dept.cell(row=i, column=j, value=val)
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
            cell.border = thin_border

    # 合計行
    total_row = len(departments) + 2
    ws_dept.cell(row=total_row, column=1, value="合計").font = Font(bold=True)
    ws_dept.cell(row=total_row, column=1).border = thin_border
    for j in range(2, 2 + len(months)):
        val = sum(ws_dept.cell(row=r, column=j).value for r in range(2, total_row))
        cell = ws_dept.cell(row=total_row, column=j, value=val)
        cell.number_format = "#,##0"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border

    # --- シート3: KPI ---
    ws_kpi = wb.create_sheet("KPI")
    kpi_items = [
        ("新規受注件数", 80, 120, "件"),
        ("既存顧客売上比率", 60, 75, "%"),
        ("顧客単価（千円）", 400, 600, "千円"),
        ("従業員数", 150, 180, "人"),
        ("一人当たり売上（千円）", 300, 380, "千円"),
    ]

    ws_kpi.cell(row=1, column=1, value="KPI項目").font = header_font
    ws_kpi.cell(row=1, column=1).fill = header_fill
    ws_kpi.cell(row=1, column=1).border = thin_border
    ws_kpi.column_dimensions["A"].width = 26

    for j, month in enumerate(months, start=2):
        cell = ws_kpi.cell(row=1, column=j, value=month)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws_kpi.column_dimensions[get_column_letter(j)].width = 14

    for i, (kpi_name, low, high, unit) in enumerate(kpi_items, start=2):
        ws_kpi.cell(row=i, column=1, value=f"{kpi_name}（{unit}）").border = thin_border
        for j in range(2, 2 + len(months)):
            growth_factor = 1.0 + 0.01 * (j - 2)
            val = round(random.uniform(low, high) * growth_factor, 1)
            if unit in ("件", "人"):
                val = int(val)
            cell = ws_kpi.cell(row=i, column=j, value=val)
            cell.number_format = "#,##0.0" if unit not in ("件", "人") else "#,##0"
            cell.alignment = Alignment(horizontal="right")
            cell.border = thin_border

    import os

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    print(f"サンプルデータを生成しました: {filepath}")


if __name__ == "__main__":
    generate_sample_excel()
